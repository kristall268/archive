"""
Модуль для сохранения/загрузки данных и экспорта в Excel
"""
import json
import os
from pathlib import Path
from typing import List, Optional
from datetime import datetime
from models import Task, TaskManager


class DataStorage:
    """Класс для работы с хранением данных"""

    def __init__(self, filename: str = "project_data.json"):
        self.filename = filename
        self.filepath = Path.cwd() / filename

    def save_tasks(self, tasks: List[Task]) -> bool:
        """Сохранить задачи в JSON файл"""
        try:
            data = {
                "version": "1.0",
                "saved_at": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
                "tasks": [task.to_dict() for task in tasks]
            }

            with open(self.filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            return True
        except Exception as e:
            print(f"Ошибка при сохранении: {e}")
            return False

    def load_tasks(self) -> Optional[List[Task]]:
        """Загрузить задачи из JSON файла"""
        if not self.filepath.exists():
            return None

        try:
            with open(self.filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)

            tasks = [Task.from_dict(task_data) for task_data in data.get("tasks", [])]
            return tasks
        except Exception as e:
            print(f"Ошибка при загрузке: {e}")
            return None

    def file_exists(self) -> bool:
        """Проверить существование файла данных"""
        return self.filepath.exists()


class ExcelExporter:
    """Класс для экспорта данных в Excel"""

    @staticmethod
    def export_to_excel(tasks: List[Task], filename: Optional[str] = None) -> tuple[bool, str]:
        """
        Экспортировать задачи в Excel файл
        
        Returns:
            tuple: (success: bool, message: str)
        """
        try:
            # Проверяем наличие openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                return False, "Библиотека openpyxl не установлена. Выполните: pip install openpyxl"

            # Определяем имя файла
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"project_export_{timestamp}.xlsx"

            # Создаем книгу и лист
            wb = Workbook()
            ws = wb.active
            ws.title = "Задачи проекта"

            # Стили
            header_font = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="3B8ED0", end_color="3B8ED0", fill_type="solid")
            cell_font = Font(name='Segoe UI', size=10)
            border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Заголовки
            headers = ["ID", "Объект", "Дата начала", "Дата окончания", 
                      "Длительность", "Зависимости", "Тип зависимости"]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

            # Ширина колонок
            column_widths = [18, 30, 15, 15, 12, 35, 20]
            for col, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64 + col)].width = width

            # Данные
            for row_idx, task in enumerate(tasks, start=2):
                # Формируем текст зависимостей
                deps_text = "\n".join(task.dependencies) if task.dependencies else "Нет"

                row_data = [
                    task.id,
                    task.object,
                    task.start_date,
                    task.end_date,
                    f"{task.duration} дней",
                    deps_text,
                    task.type if task.type else "--"
                ]

                for col, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.font = cell_font
                    cell.border = border
                    
                    # Выравнивание
                    if col in [1, 3, 4, 5, 7]:  # ID, даты, длительность, тип
                        cell.alignment = center_align
                    else:  # Объект, зависимости
                        cell.alignment = left_align

                # Автоматическая высота строки для зависимостей
                if task.dependencies and len(task.dependencies) > 1:
                    ws.row_dimensions[row_idx].height = 15 * len(task.dependencies)

            # Добавляем информацию о проекте
            info_row = len(tasks) + 3
            ws.cell(row=info_row, column=1, value="Дата экспорта:").font = Font(bold=True)
            ws.cell(row=info_row, column=2, value=datetime.now().strftime("%d.%m.%Y %H:%M:%S"))
            
            ws.cell(row=info_row + 1, column=1, value="Всего задач:").font = Font(bold=True)
            ws.cell(row=info_row + 1, column=2, value=len(tasks))

            # Сохраняем файл
            wb.save(filename)
            return True, f"Данные экспортированы в файл: {filename}"

        except Exception as e:
            return False, f"Ошибка при экспорте: {str(e)}"


class AutoSaveManager:
    """Менеджер автоматического сохранения"""

    def __init__(self, task_manager: TaskManager, storage: DataStorage, 
                 interval_ms: int = 30000):  # 30 секунд
        self.task_manager = task_manager
        self.storage = storage
        self.interval_ms = interval_ms
        self.auto_save_enabled = True
        self.parent = None
        self.save_job = None

    def start(self, parent):
        """Запустить автосохранение"""
        self.parent = parent
        self._schedule_save()

    def stop(self):
        """Остановить автосохранение"""
        if self.save_job:
            try:
                self.parent.after_cancel(self.save_job)
            except:
                pass
        self.save_job = None

    def _schedule_save(self):
        """Запланировать следующее сохранение"""
        if not self.auto_save_enabled or not self.parent:
            return

        self.save_job = self.parent.after(self.interval_ms, self._perform_save)

    def _perform_save(self):
        """Выполнить сохранение"""
        if self.auto_save_enabled:
            tasks = self.task_manager.get_all_tasks()
            self.storage.save_tasks(tasks)
            # Планируем следующее сохранение
            self._schedule_save()

    def save_now(self):
        """Сохранить немедленно"""
        tasks = self.task_manager.get_all_tasks()
        return self.storage.save_tasks(tasks)

    def toggle_auto_save(self, enabled: bool):
        """Включить/выключить автосохранение"""
        self.auto_save_enabled = enabled
        if enabled and self.parent:
            self._schedule_save()
        elif self.save_job:
            self.stop()